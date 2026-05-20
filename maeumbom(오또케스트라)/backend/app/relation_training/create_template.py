"""
Excel 템플릿 파일 생성 스크립트
시나리오 데이터 작성을 위한 템플릿 생성
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path


def create_template():
    """템플릿 Excel 파일 생성"""
    wb = Workbook()
    
    # 기본 시트 삭제
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # ========================================================================
    # 시트 1: scenarios
    # ========================================================================
    ws_scenarios = wb.create_sheet('scenarios', 0)
    
    # 헤더
    headers_scenarios = ['scenario_id', 'title', 'target_type', 'category']
    ws_scenarios.append(headers_scenarios)
    
    # 헤더 스타일
    for cell in ws_scenarios[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # 예시 데이터
    ws_scenarios.append([
        1,
        '부모님과의 대화',
        'parent',
        'TRAINING'
    ])
    
    # 설명 추가
    ws_scenarios.append([])
    ws_scenarios.append(['# target_type: parent, friend, partner, child, colleague'])
    ws_scenarios.append(['# category: TRAINING (관계 개선 훈련), DRAMA (공감 드라마)'])
    
    # 컬럼 너비 조정
    ws_scenarios.column_dimensions['A'].width = 15
    ws_scenarios.column_dimensions['B'].width = 30
    ws_scenarios.column_dimensions['C'].width = 15
    ws_scenarios.column_dimensions['D'].width = 15
    
    # ========================================================================
    # 시트 2: nodes
    # ========================================================================
    ws_nodes = wb.create_sheet('nodes', 1)
    
    # 헤더
    headers_nodes = ['scenario_id', 'step_level', 'situation_text', 'image_url']
    ws_nodes.append(headers_nodes)
    
    # 헤더 스타일
    for cell in ws_nodes[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # 예시 데이터
    ws_nodes.append([
        1,
        1,
        '부모님이 당신의 진로에 대해 걱정하며 이야기를 꺼내십니다.',
        ''
    ])
    ws_nodes.append([
        1,
        2,
        '부모님이 당신의 이야기를 진지하게 듣고 계십니다.',
        ''
    ])
    
    # 설명 추가
    ws_nodes.append([])
    ws_nodes.append(['# step_level: 1부터 시작 (1, 2, 3, 4...)'])
    ws_nodes.append(['# image_url: 선택사항 (비워두면 NULL)'])
    
    # 컬럼 너비 조정
    ws_nodes.column_dimensions['A'].width = 15
    ws_nodes.column_dimensions['B'].width = 12
    ws_nodes.column_dimensions['C'].width = 60
    ws_nodes.column_dimensions['D'].width = 30
    
    # ========================================================================
    # 시트 3: options
    # ========================================================================
    ws_options = wb.create_sheet('options', 2)
    
    # 헤더
    headers_options = ['scenario_id', 'node_step', 'option_code', 'option_text', 'next_step', 'result_code']
    ws_options.append(headers_options)
    
    # 헤더 스타일
    for cell in ws_options[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # 예시 데이터
    ws_options.append([
        1,
        1,
        'A',
        '부모님의 걱정을 이해하고 솔직하게 내 상황을 설명한다',
        2,
        ''
    ])
    ws_options.append([
        1,
        1,
        'B',
        '괜찮다고만 말하고 대화를 피한다',
        '',
        'FAIL'
    ])
    ws_options.append([
        1,
        2,
        'A',
        '내 진짜 마음과 고민을 솔직하게 털어놓는다',
        '',
        'SUCCESS'
    ])
    ws_options.append([
        1,
        2,
        'B',
        '부모님이 원하는 답을 하며 대화를 마무리한다',
        '',
        'PARTIAL'
    ])
    
    # 설명 추가
    ws_options.append([])
    ws_options.append(['# node_step: 이 선택지가 속한 노드의 step_level'])
    ws_options.append(['# option_code: A, B, C, D... (각 노드에서 고유해야 함)'])
    ws_options.append(['# next_step: 다음 노드의 step_level (빈칸이면 결과로)'])
    ws_options.append(['# result_code: next_step이 빈칸일 때 필수 (결과 코드)'])
    
    # 컬럼 너비 조정
    ws_options.column_dimensions['A'].width = 15
    ws_options.column_dimensions['B'].width = 12
    ws_options.column_dimensions['C'].width = 12
    ws_options.column_dimensions['D'].width = 60
    ws_options.column_dimensions['E'].width = 12
    ws_options.column_dimensions['F'].width = 15
    
    # ========================================================================
    # 시트 4: results
    # ========================================================================
    ws_results = wb.create_sheet('results', 3)
    
    # 헤더
    headers_results = ['scenario_id', 'result_code', 'display_title', 'analysis_text', 'atmosphere_image_type', 'score']
    ws_results.append(headers_results)
    
    # 헤더 스타일
    for cell in ws_results[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # 예시 데이터
    ws_results.append([
        1,
        'SUCCESS',
        '성공적인 대화',
        '부모님과 솔직하고 진솔한 대화를 나누셨습니다. 서로의 입장을 이해하고 공감하는 시간이 되었습니다.',
        'positive',
        85
    ])
    ws_results.append([
        1,
        'PARTIAL',
        '부분적 성공',
        '대화는 나눴지만 완전히 마음을 열지는 못했습니다. 다음에는 좀 더 솔직하게 이야기해보세요.',
        'neutral',
        60
    ])
    ws_results.append([
        1,
        'FAIL',
        '대화 실패',
        '대화가 원활하게 이루어지지 않았습니다. 부모님의 마음을 이해하고 경청하는 자세가 필요합니다.',
        'negative',
        30
    ])
    
    # 설명 추가
    ws_results.append([])
    ws_results.append(['# result_code: 고유한 결과 코드 (SUCCESS, FAIL, PARTIAL...)'])
    ws_results.append(['# atmosphere_image_type: positive, negative, neutral'])
    ws_results.append(['# score: 0-100 점수 (선택사항)'])
    
    # 컬럼 너비 조정
    ws_results.column_dimensions['A'].width = 15
    ws_results.column_dimensions['B'].width = 15
    ws_results.column_dimensions['C'].width = 20
    ws_results.column_dimensions['D'].width = 60
    ws_results.column_dimensions['E'].width = 20
    ws_results.column_dimensions['F'].width = 10
    
    # 파일 저장
    data_dir = Path(__file__).parent / 'data'
    data_dir.mkdir(exist_ok=True)  # data 폴더가 없으면 생성
    
    template_path = data_dir / 'template.xlsx'
    wb.save(template_path)
    print(f"✅ 템플릿 파일 생성 완료: {template_path}")


if __name__ == '__main__':
    create_template()

