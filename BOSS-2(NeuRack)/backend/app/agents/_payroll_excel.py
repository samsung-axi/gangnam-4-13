"""급여명세서 Excel 템플릿 선택·채움 모듈."""
from __future__ import annotations

import io
from pathlib import Path
from typing import TypedDict

from openpyxl import load_workbook

TEMPLATE_DIR = Path(__file__).parent / "_payroll_templates"


class PayrollData(TypedDict, total=False):
    name: str
    employee_id: str
    department: str
    position: str
    pay_date: str  # YYYY-MM-DD
    employment_type: str  # 초단시간 | 시급제 | 월급제
    has_allowances: bool
    hourly_rate: int
    hours_worked: float
    base_pay: int
    overtime_hours: float
    overtime_pay: int
    night_hours: float
    night_pay: int
    holiday_hours: float
    holiday_pay: int
    meal_allowance: int
    family_allowance: int
    income_tax: int
    national_pension: int
    health_insurance: int
    ltc_insurance: int
    employment_insurance: int
    other_deductions: dict
    total_pay: int
    total_deductions: int
    net_pay: int


def select_template(employment_type: str, has_allowances: bool) -> Path:
    emp = employment_type.strip()
    if "초단시간" in emp:
        fname = "template_3_시급_수당.xlsx" if has_allowances else "template_1_초단시간_단순.xlsx"
    elif "월급" in emp:
        fname = "template_5_월급_수당.xlsx"
    else:
        fname = "template_3_시급_수당.xlsx" if has_allowances else "template_2_시급_단순.xlsx"
    return TEMPLATE_DIR / fname


def _set(ws, coord: str, value) -> None:
    """병합 셀 여부와 무관하게 좌상단 셀에 값 설정."""
    cell = ws[coord]
    if hasattr(cell, "coordinate"):
        cell.value = value


def _fill_template_1(ws, data: PayrollData) -> None:
    """초단시간, 수당공제 없음 (13r x 6c)."""
    if data.get("pay_date"):
        _set(ws, "A3", f"지급일 : {data['pay_date']} ")
    if data.get("name"):
        _set(ws, "B4", data["name"])
    if data.get("base_pay"):
        _set(ws, "D9", data["base_pay"])
    hr = data.get("hourly_rate")
    hw = data.get("hours_worked")
    if hr and hw:
        _set(ws, "D10", f"({hw}시간x{hr:,}원)")
    if data.get("total_pay"):
        _set(ws, "D12", data["total_pay"])


def _fill_template_2(ws, data: PayrollData) -> None:
    """비초단시간, 수당공제 없음 (13r x 6c) — template_1 과 동일 구조."""
    _fill_template_1(ws, data)


def _fill_template_3(ws, data: PayrollData) -> None:
    """시급/일급, 수당공제 있음 (20r x 8c)."""
    if data.get("pay_date"):
        _set(ws, "A3", f"지급일 : {data['pay_date']} ")
    if data.get("name"):
        _set(ws, "B4", data["name"])

    # 지급 내역
    if data.get("base_pay"):
        _set(ws, "B9", data["base_pay"])
    if data.get("meal_allowance"):
        _set(ws, "B10", data["meal_allowance"])

    # 공제 내역
    if data.get("income_tax") is not None:
        _set(ws, "F9", data["income_tax"])
    if data.get("national_pension"):
        _set(ws, "F10", data["national_pension"])
    if data.get("health_insurance"):
        _set(ws, "F11", data["health_insurance"])
    if data.get("ltc_insurance"):
        _set(ws, "F12", data["ltc_insurance"])
    if data.get("employment_insurance"):
        _set(ws, "F13", data["employment_insurance"])

    # 합계
    if data.get("total_pay"):
        _set(ws, "B14", data["total_pay"])
    if data.get("total_deductions"):
        _set(ws, "F14", data["total_deductions"])
    if data.get("net_pay"):
        _set(ws, "F15", data["net_pay"])

    # 계산방법 행 (r19)
    hr = data.get("hourly_rate")
    hw = data.get("hours_worked")
    if hr and hw:
        _set(ws, "A19", "기본급")
        _set(ws, "C19", f" {hw}시간 x {hr:,}원")
        if data.get("base_pay"):
            _set(ws, "H19", data["base_pay"])


def _fill_template_5(ws, data: PayrollData) -> None:
    """월급제, 수당공제 있음 (29r x 10c)."""
    if data.get("pay_date"):
        _set(ws, "A3", f"지급일 : {data['pay_date']} ")
    if data.get("name"):
        _set(ws, "C4", data["name"])
    if data.get("employee_id"):
        _set(ws, "H4", data["employee_id"])
    if data.get("department"):
        _set(ws, "C5", data["department"])
    if data.get("position"):
        _set(ws, "H5", data["position"])

    # 지급 내역
    if data.get("base_pay"):
        _set(ws, "D10", data["base_pay"])
    if data.get("overtime_pay"):
        _set(ws, "D11", data["overtime_pay"])
    if data.get("night_pay"):
        _set(ws, "D12", data["night_pay"])
    if data.get("holiday_pay"):
        _set(ws, "D13", data["holiday_pay"])
    if data.get("family_allowance"):
        _set(ws, "D14", data["family_allowance"])
    if data.get("meal_allowance"):
        _set(ws, "D15", data["meal_allowance"])

    # 공제 내역
    if data.get("income_tax") is not None:
        _set(ws, "H10", data["income_tax"])
    if data.get("national_pension"):
        _set(ws, "H11", data["national_pension"])
    if data.get("employment_insurance"):
        _set(ws, "H12", data["employment_insurance"])
    if data.get("health_insurance"):
        _set(ws, "H13", data["health_insurance"])
    if data.get("ltc_insurance"):
        _set(ws, "H14", data["ltc_insurance"])

    # 합계
    if data.get("total_pay"):
        _set(ws, "D19", data["total_pay"])
    if data.get("total_deductions"):
        _set(ws, "H19", data["total_deductions"])
    if data.get("net_pay"):
        _set(ws, "H20", data["net_pay"])

    # 계산방법 (r24~)
    calc_row = 24
    items: list[tuple[str, str, int]] = []
    if data.get("overtime_pay") and data.get("overtime_hours"):
        items.append(("연장근로수당", f"{data['overtime_hours']}시간×통상시급×1.5", data["overtime_pay"]))
    if data.get("night_pay") and data.get("night_hours"):
        items.append(("야간근로수당", f"{data['night_hours']}시간×통상시급×0.5", data["night_pay"]))
    if data.get("holiday_pay") and data.get("holiday_hours"):
        items.append(("휴일근로수당", f"{data['holiday_hours']}시간×통상시급×1.5", data["holiday_pay"]))
    for i, (label, formula, amount) in enumerate(items):
        _set(ws, f"A{calc_row + i}", label)
        _set(ws, f"E{calc_row + i}", formula)
        _set(ws, f"J{calc_row + i}", amount)


_FILLERS = {
    "template_1_초단시간_단순.xlsx": _fill_template_1,
    "template_2_시급_단순.xlsx": _fill_template_2,
    "template_3_시급_수당.xlsx": _fill_template_3,
    "template_5_월급_수당.xlsx": _fill_template_5,
}


def generate_payroll_excel(data: PayrollData) -> tuple[bytes, str]:
    """PayrollData → (Excel bytes, 파일명)."""
    emp = data.get("employment_type", "시급제")
    has_allowances = bool(data.get("has_allowances", True))
    template_path = select_template(emp, has_allowances)

    wb = load_workbook(template_path)
    ws = wb.worksheets[0]

    filler = _FILLERS.get(template_path.name)
    if filler:
        filler(ws, data)

    buf = io.BytesIO()
    wb.save(buf)

    name = data.get("name", "직원")
    pay_date = (data.get("pay_date") or "")[:7].replace("-", "")
    filename = f"{name}_{pay_date}_급여명세서.xlsx"
    return buf.getvalue(), filename
