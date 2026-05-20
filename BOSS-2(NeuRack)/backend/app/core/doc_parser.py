"""업로드된 서류의 텍스트 추출기.

지원 포맷:
  - PDF  (PyMuPDF, 텍스트 기반)
  - DOCX / DOC (python-docx)
  - TXT  (plain text, utf-8 / cp949 auto)
  - RTF  (striprtf)
  - XLSX (openpyxl → 시트별 markdown 테이블)
  - CSV  (stdlib csv → markdown 테이블)
  - 이미지(JPG/PNG/WEBP/BMP/TIFF/GIF/HEIC) — OpenAI gpt-4o vision OCR
    · HEIC 는 pillow-heif 로 JPEG 변환 후 기존 OCR 파이프로 주입.

최상위 진입점 `parse_file` 은 async (이미지 OCR 이 async 이기 때문).
분석·저장 로직과 분리된 순수 변환 함수만 노출.
"""

from __future__ import annotations

import csv
import io
import logging
from typing import Iterable

from fastapi import HTTPException

log = logging.getLogger(__name__)

SUPPORTED_EXTS: tuple[str, ...] = (
    "pdf", "docx", "doc",
    "txt", "rtf",
    "xlsx", "csv",
)
IMAGE_EXTS: tuple[str, ...] = ("jpg", "jpeg", "png", "webp", "bmp", "tiff", "gif", "heic", "heif")


def _ext_of(filename: str) -> str:
    return filename.rsplit(".", 1)[-1].lower() if "." in filename else ""


# ─────────────────────────── 바이너리 포맷 ───────────────────────────

def parse_pdf(file_bytes: bytes) -> str:
    try:
        import fitz  # PyMuPDF
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="PDF 파싱 모듈(pymupdf) 이 설치되지 않았습니다.",
        )

    doc = fitz.open(stream=file_bytes, filetype="pdf")
    try:
        text = "\n".join(page.get_text() for page in doc).strip()
    finally:
        doc.close()

    if len(text) < 50:
        raise HTTPException(
            status_code=422,
            detail="PDF 에서 텍스트를 추출하지 못했습니다. 스캔 PDF 는 이미지 형태(JPG/PNG) 로 올려주시면 OCR 로 처리됩니다.",
        )
    return text


def parse_docx(file_bytes: bytes) -> str:
    try:
        from docx import Document
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="DOCX 파싱 모듈(python-docx) 이 설치되지 않았습니다.",
        )
    doc = Document(io.BytesIO(file_bytes))
    text = "\n".join(p.text for p in doc.paragraphs).strip()
    if not text:
        raise HTTPException(status_code=422, detail="DOCX 에서 텍스트를 추출하지 못했습니다.")
    return text


# ─────────────────────────── 텍스트 계열 ───────────────────────────

def parse_txt(file_bytes: bytes) -> str:
    """UTF-8 → CP949 → Latin-1 순으로 디코딩 시도."""
    for enc in ("utf-8", "utf-8-sig", "cp949", "euc-kr", "latin-1"):
        try:
            text = file_bytes.decode(enc).strip()
            if text:
                return text
        except UnicodeDecodeError:
            continue
    raise HTTPException(status_code=422, detail="TXT 파일 인코딩을 인식하지 못했습니다.")


def parse_rtf(file_bytes: bytes) -> str:
    try:
        from striprtf.striprtf import rtf_to_text
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="RTF 파싱 모듈(striprtf) 이 설치되지 않았습니다.",
        )
    # striprtf 는 str 을 받음 → 최대 관대하게 디코딩
    try:
        raw = file_bytes.decode("utf-8", errors="ignore")
    except Exception:
        raw = file_bytes.decode("latin-1", errors="ignore")
    text = rtf_to_text(raw, errors="ignore").strip()
    if not text:
        raise HTTPException(status_code=422, detail="RTF 에서 텍스트를 추출하지 못했습니다.")
    return text


# ─────────────────────────── 스프레드시트 ───────────────────────────

def _rows_to_markdown(rows: Iterable[list[str]]) -> str:
    rows = [r for r in rows if any((c or "").strip() for c in r)]
    if not rows:
        return ""
    header = rows[0]
    body = rows[1:] if len(rows) > 1 else []
    head_line = "| " + " | ".join((c or "").strip() for c in header) + " |"
    sep_line = "|" + "|".join("---" for _ in header) + "|"
    body_lines = [
        "| " + " | ".join((c or "").strip() for c in r) + " |"
        for r in body
    ]
    return "\n".join([head_line, sep_line, *body_lines])


def parse_xlsx(file_bytes: bytes) -> str:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="XLSX 파싱 모듈(openpyxl) 이 설치되지 않았습니다.",
        )
    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    except Exception as exc:
        raise HTTPException(status_code=422, detail=f"XLSX 열기 실패: {str(exc)[:150]}")

    chunks: list[str] = []
    for sheet in wb.worksheets:
        rows: list[list[str]] = []
        for raw in sheet.iter_rows(values_only=True):
            rows.append(["" if v is None else str(v) for v in raw])
        md = _rows_to_markdown(rows)
        if md:
            chunks.append(f"## [시트] {sheet.title}\n\n{md}")
    wb.close()
    text = "\n\n".join(chunks).strip()
    if not text:
        raise HTTPException(status_code=422, detail="XLSX 에 추출할 내용이 없습니다.")
    return text


def parse_csv(file_bytes: bytes) -> str:
    # TXT 와 동일한 인코딩 탐색
    raw: str | None = None
    for enc in ("utf-8-sig", "utf-8", "cp949", "euc-kr", "latin-1"):
        try:
            raw = file_bytes.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if raw is None:
        raise HTTPException(status_code=422, detail="CSV 파일 인코딩을 인식하지 못했습니다.")

    # 구분자 자동 감지 (comma / semicolon / tab)
    sample = raw[:2048]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(io.StringIO(raw), dialect)
    rows = [list(r) for r in reader]
    md = _rows_to_markdown(rows)
    if not md:
        raise HTTPException(status_code=422, detail="CSV 에서 데이터를 추출하지 못했습니다.")
    return md


# ─────────────────────────── 이미지 (OCR) ───────────────────────────

def _heic_to_jpeg(file_bytes: bytes) -> bytes:
    """HEIC/HEIF → JPEG 변환. 실패 시 HTTPException."""
    try:
        import pillow_heif  # type: ignore
        from PIL import Image  # type: ignore
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="HEIC 변환 모듈(pillow-heif) 이 설치되지 않았습니다. JPG/PNG 로 변환 후 올려주세요.",
        )
    try:
        heif = pillow_heif.read_heif(file_bytes)
        img = Image.frombytes(heif.mode, heif.size, heif.data, "raw")
        buf = io.BytesIO()
        img.convert("RGB").save(buf, format="JPEG", quality=92)
        return buf.getvalue()
    except Exception as exc:
        raise HTTPException(status_code=422, detail=f"HEIC 변환 실패: {str(exc)[:150]}")


# ─────────────────────────── dispatcher ───────────────────────────

async def parse_file(file_bytes: bytes, filename: str) -> str:
    ext = _ext_of(filename)
    if ext == "pdf":
        return parse_pdf(file_bytes)
    if ext in ("docx", "doc"):
        return parse_docx(file_bytes)
    if ext == "txt":
        return parse_txt(file_bytes)
    if ext == "rtf":
        return parse_rtf(file_bytes)
    if ext == "xlsx":
        return parse_xlsx(file_bytes)
    if ext == "csv":
        return parse_csv(file_bytes)
    if ext in IMAGE_EXTS:
        from app.core.ocr import extract_text_from_image
        if ext in ("heic", "heif"):
            # JPEG 로 변환한 뒤 OCR — filename 도 .jpg 로 정규화해 mime 판정 정확히
            jpeg_bytes = _heic_to_jpeg(file_bytes)
            stem = filename.rsplit(".", 1)[0] if "." in filename else filename
            return await extract_text_from_image(jpeg_bytes, f"{stem}.jpg")
        return await extract_text_from_image(file_bytes, filename)
    raise HTTPException(
        status_code=415,
        detail=f"지원하지 않는 파일 형식: .{ext}. PDF·DOCX·TXT·RTF·XLSX·CSV·이미지(JPG/PNG/WEBP/BMP/TIFF/GIF/HEIC) 만 업로드할 수 있어요.",
    )
